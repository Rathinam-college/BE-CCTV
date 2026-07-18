from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.views import APIView
from django.core.management import call_command
from django.core import serializers
from django.apps import apps
from django.db.models.functions import TruncMonth
import json
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import (
    Camera, NVR, Biometric, Barrier, NetworkSwitch, ActivityLog, 
    CameraRemark, NVRRemark, BiometricRemark, SwitchRemark,
    CameraRelocation, NVRRelocation, BiometricRelocation, SwitchRelocation,
    GlobalSiteConfig,
    Rack, RackRemark, RackRelocation, Division, Brand
)
from .serializers import (
    CameraSerializer, NVRSerializer, BiometricSerializer, BarrierSerializer, 
    NetworkSwitchSerializer, ActivityLogSerializer, CameraRemarkSerializer, 
    NVRRemarkSerializer, BiometricRemarkSerializer, SwitchRemarkSerializer,
    CameraRelocationSerializer, NVRRelocationSerializer, BiometricRelocationSerializer, SwitchRelocationSerializer,
    GlobalSiteConfigSerializer,
    RackSerializer, RackRemarkSerializer, RackRelocationSerializer, DivisionSerializer, BrandSerializer
)

from django.db import transaction
from django.db.models import Count
import openpyxl
import io

def log_activity(user, action, page, details='', request=None):
    ip = ''
    if request:
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
    
    ActivityLog.objects.create(
        user=user,
        action=action,
        page=page,
        details=details,
        ipAddress=ip
    )

class DatabaseBackupView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'Super Admin':
            return Response({'error': 'Unauthorized'}, status=403)
            
        action_type = request.query_params.get('action')
        if action_type == 'get_months':
            months = set()
            try:
                # Query distinct months from ActivityLog (which generally covers everything) or Tickets/Cameras
                # ActivityLog is a good proxy for activity
                logs = ActivityLog.objects.annotate(month=TruncMonth('timestamp')).values('month').distinct()
                for log in logs:
                    if log['month']:
                        months.add(log['month'].strftime('%Y-%m'))
                        
                from maintenance.models import Ticket
                tickets = Ticket.objects.annotate(month=TruncMonth('createdAt')).values('month').distinct()
                for t in tickets:
                    if t['month']:
                        months.add(t['month'].strftime('%Y-%m'))
            except Exception:
                pass
            return Response({'months': sorted(list(months), reverse=True)})
        
        format_type = request.query_params.get('format', 'json')
        
        if format_type == 'sql':
            from django.conf import settings
            db = settings.DATABASES['default']
            
            if 'sqlite' in db['ENGINE'].lower():
                import sqlite3
                import io
                db_path = db['NAME']
                try:
                    conn = sqlite3.connect(db_path)
                    output = io.StringIO()
                    for line in conn.iterdump():
                        output.write('%s\n' % line)
                    sql_data = output.getvalue().encode('utf-8')
                    conn.close()
                    response = HttpResponse(sql_data, content_type='application/sql')
                    response['Content-Disposition'] = 'attachment; filename="backup.sql"'
                    return response
                except Exception as e:
                    return Response({'error': f"SQLite dump failed: {str(e)}"}, status=400)
            
            import subprocess
            import os
            
            env = os.environ.copy()
            if db.get('PASSWORD'):
                env['PGPASSWORD'] = str(db['PASSWORD'])
                
            cmd = [
                'pg_dump',
                '-h', str(db.get('HOST', 'localhost')),
                '-U', str(db.get('USER', 'postgres')),
                '-d', str(db.get('NAME', 'postgres')),
                '--clean', '--if-exists', '--inserts', '--no-owner', '--no-privileges'
            ]
            if db.get('PORT'):
                cmd.extend(['-p', str(db['PORT'])])
                
            try:
                # Add schema-only for now, but user wants 'full data'. 
                # pg_dump with --inserts will dump schema + data.
                result = subprocess.run(cmd, env=env, capture_output=True)
                if result.returncode != 0:
                    return Response({'error': f"pg_dump failed: {result.stderr.decode('utf-8', errors='ignore')}"}, status=400)
                
                response = HttpResponse(result.stdout, content_type='application/sql')
                response['Content-Disposition'] = 'attachment; filename="backup.sql"'
                return response
            except FileNotFoundError:
                return Response({'error': 'pg_dump command not found on the server. Please ensure PostgreSQL client tools are installed and in the system PATH.'}, status=400)
            except Exception as e:
                return Response({'error': f"SQL backup failed: {str(e)}"}, status=400)
        
        models_to_dump = request.query_params.get('models')
        month_filter = request.query_params.get('month')
        
        args = []
        if models_to_dump:
            args = models_to_dump.split(',')
        else:
            args = ['cctv.camera', 'cctv.nvr', 'cctv.biometric', 'cctv.networkswitch', 'cctv.rack', 'maintenance.ticket', 'maintenance.project', 'users.user', 'cctv.activitylog', 'cctv.masterlocation', 'cctv.block', 'cctv.floor', 'cctv.room']
            
        try:
            objects_to_serialize = []
            for model_str in args:
                try:
                    app_label, model_name = model_str.split('.')
                    model = apps.get_model(app_label, model_name)
                    qs = model.objects.all()
                    
                    if month_filter:
                        year, mo = month_filter.split('-')
                        # Check for date fields
                        date_fields = [f.name for f in model._meta.fields if f.get_internal_type() in ['DateTimeField', 'DateField']]
                        if 'createdAt' in date_fields:
                            qs = qs.filter(createdAt__year=year, createdAt__month=mo)
                        elif 'timestamp' in date_fields:
                            qs = qs.filter(timestamp__year=year, timestamp__month=mo)
                        elif 'date_joined' in date_fields:
                            qs = qs.filter(date_joined__year=year, date_joined__month=mo)
                            
                    objects_to_serialize.extend(list(qs))
                except Exception as inner_e:
                    print(f"Skipping {model_str}: {inner_e}")
                    
            data = serializers.serialize('json', objects_to_serialize)
            
            response = HttpResponse(data, content_type='application/json')
            response['Content-Disposition'] = 'attachment; filename="backup.json"'
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def delete(self, request):
        if request.user.role != 'Super Admin':
            return Response({'error': 'Unauthorized'}, status=403)
            
        models_to_delete = request.query_params.get('models')
        if not models_to_delete:
            return Response({'error': 'No models specified for deletion. Please select datasets to clear.'}, status=400)
            
        models_list = models_to_delete.split(',')
        
        try:
            for model_str in models_list:
                try:
                    app_label, model_name = model_str.split('.')
                    model = apps.get_model(app_label, model_name)
                    model.objects.all().delete()
                except Exception as inner_e:
                    print(f"Skipping deletion for {model_str}: {inner_e}")
            return Response({'status': 'Selected data cleared successfully'})
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def post(self, request):
        if request.user.role != 'Super Admin':
            return Response({'error': 'Unauthorized'}, status=403)
            
        file_obj = request.FILES.get('backup_file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=400)
            
        try:
            if file_obj.name.endswith('.sql'):
                from django.conf import settings
                db = settings.DATABASES['default']
                
                if 'sqlite' in db['ENGINE'].lower():
                    import sqlite3
                    sql_content = file_obj.read().decode('utf-8', errors='ignore')
                    db_path = db['NAME']
                    try:
                        conn = sqlite3.connect(db_path)
                        cursor = conn.cursor()
                        cursor.execute("PRAGMA foreign_keys = OFF;")
                        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%';")
                        tables = [row[0] for row in cursor.fetchall()]
                        for table in tables:
                            cursor.execute(f"DROP TABLE IF EXISTS \"{table}\";")
                        cursor.executescript(sql_content)
                        conn.commit()
                        conn.close()
                        return Response({'status': 'SQLite database restored successfully!'})
                    except Exception as e:
                        return Response({'error': f"SQLite restore failed: {str(e)}"}, status=400)
                
                import subprocess
                import os
                import tempfile
                
                with tempfile.NamedTemporaryFile(delete=False, suffix='.sql') as tmp:
                    # Read all, replace old Ticket table names, write to tmp
                    sql_content = file_obj.read().decode('utf-8', errors='ignore')
                    sql_content = sql_content.replace('public.cctv_ticket', 'public.maintenance_ticket')
                    sql_content = sql_content.replace('INSERT INTO cctv_ticket', 'INSERT INTO maintenance_ticket')
                    sql_content = sql_content.replace('public.cctv_ticketremark', 'public.maintenance_ticketremark')
                    sql_content = sql_content.replace('INSERT INTO cctv_ticketremark', 'INSERT INTO maintenance_ticketremark')
                    sql_content = sql_content.replace('public.cctv_ticketdocument', 'public.maintenance_ticketdocument')
                    sql_content = sql_content.replace('INSERT INTO cctv_ticketdocument', 'INSERT INTO maintenance_ticketdocument')
                    sql_content = sql_content.replace('public.cctv_ticketcompletedimage', 'public.maintenance_ticketcompletedimage')
                    sql_content = sql_content.replace('INSERT INTO cctv_ticketcompletedimage', 'INSERT INTO maintenance_ticketcompletedimage')
                    sql_content = sql_content.replace('"cctv_ticket"', '"maintenance_ticket"')
                    sql_content = sql_content.replace('"cctv_ticketremark"', '"maintenance_ticketremark"')
                    sql_content = sql_content.replace('"cctv_ticketdocument"', '"maintenance_ticketdocument"')
                    sql_content = sql_content.replace('"cctv_ticketcompletedimage"', '"maintenance_ticketcompletedimage"')
                    tmp.write(sql_content.encode('utf-8'))
                    tmp_path = tmp.name
                    
                env = os.environ.copy()
                if db.get('PASSWORD'):
                    env['PGPASSWORD'] = str(db['PASSWORD'])
                    
                cmd = [
                    'psql', 
                    '-h', str(db.get('HOST', 'localhost')), 
                    '-U', str(db.get('USER', 'postgres')), 
                    '-d', str(db.get('NAME', 'postgres')), 
                    '-f', tmp_path
                ]
                if db.get('PORT'):
                    cmd.extend(['-p', str(db['PORT'])])
                
                try:
                    result = subprocess.run(cmd, env=env, capture_output=True, text=True)
                    os.unlink(tmp_path)
                    if result.returncode != 0:
                        return Response({'error': f"SQL Error: {result.stderr}"}, status=400)
                    return Response({'status': 'SQL data imported successfully!'})
                except FileNotFoundError:
                    os.unlink(tmp_path)
                    return Response({'error': 'psql command not found on the server. Please ensure PostgreSQL client tools are installed in the backend environment.'}, status=400)
            else:
                data = file_obj.read().decode('utf-8', errors='ignore')
                # Map old JSON model references to new maintenance app
                data = data.replace('"model": "cctv.ticket"', '"model": "maintenance.ticket"')
                data = data.replace('"model": "cctv.ticketremark"', '"model": "maintenance.ticketremark"')
                data = data.replace('"model": "cctv.ticketdocument"', '"model": "maintenance.ticketdocument"')
                data = data.replace('"model": "cctv.ticketcompletedimage"', '"model": "maintenance.ticketcompletedimage"')
                with transaction.atomic():
                    for obj in serializers.deserialize('json', data):
                        obj.save()
                return Response({'status': 'Database restored successfully!'})
        except Exception as e:
            return Response({'error': f'Failed to restore database: {str(e)}'}, status=400)

def check_can_edit(user, page):
    if user.role == 'Super Admin':
        return True
    return f"{page}:EDIT" in user.permissions

def sync_location_to_master(college, block, floor, room, brand):
    if block:
        try:
            from .models import Block, Floor, Room
            b, _ = Block.objects.get_or_create(name=block)
            if floor:
                f, _ = Floor.objects.get_or_create(name=floor, block=b)
                if room:
                    Room.objects.get_or_create(name=room, block=b, floor=f)
        except Exception as e:
            print(f"Normalized Location sync failed: {str(e)}")

class CameraViewSet(viewsets.ModelViewSet):
    queryset = Camera.objects.all().order_by('-id')
    serializer_class = CameraSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Cameras'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        response = super().create(request, *args, **kwargs)
        log_activity(request.user, 'CREATE', 'Cameras', f"Created camera: {request.data.get('name')}", request)
        return response

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Cameras'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        data = request.data
        if not isinstance(data, list):
            return Response({'message': 'Data must be a list'}, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = self.get_serializer(data=data, many=True)
        serializer.is_valid(raise_exception=True)
        self.perform_bulk_create(serializer)
        log_activity(request.user, 'CREATE', 'Cameras', f"Bulk created {len(data)} cameras", request)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_bulk_create(self, serializer):
        serializer.save()

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Cameras'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'message': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Handle both .xlsx and .csv if needed, but primarily .xlsx
            if file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls'):
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
                rows_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    rows_data.append(dict(zip(headers, row)))
            elif file_obj.name.endswith('.csv'):
                import csv
                decoded_file = file_obj.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                rows_data = [ {k.lower(): v for k, v in row.items()} for row in reader]
            else:
                return Response({'message': 'Unsupported file format. Please upload .xlsx or .csv'}, status=status.HTTP_400_BAD_REQUEST)

            created_count = 0
            updated_count = 0
            with transaction.atomic():
                for data in rows_data:
                    # Mapping
                    name_val = data.get('device type') or data.get('devicetype') or data.get('camera type') or data.get('name') or 'Unknown Camera'
                    camera_data = {
                        'ipAddress': data.get('ip address') or data.get('ipaddress'),
                        'name': name_val,
                        'deviceType': data.get('device type') or data.get('devicetype') or data.get('camera type') or name_val,
                        'status': data.get('status') or 'Online',
                        'campusZone': data.get('campus zone') or data.get('campuszone') or 'INSIDE',
                        'block': data.get('block'),
                        'floor': data.get('floor'),
                        'brand': data.get('brand'),
                        'remarks': data.get('remarks'),
                    }
                    
                    # Reconstruct siteName and dvrNvrDetails
                    block = camera_data['block'] or ''
                    floor = camera_data['floor'] or ''
                    camera_data['siteName'] = f"{block} - {floor}" if block or floor else 'Unknown'
                    
                    gateway = data.get('ipv4 gateway') or data.get('gateway') or ''
                    subnet = data.get('subnet mask') or data.get('subnet') or ''
                    mac = data.get('mac address') or data.get('mac') or ''
                    college = data.get('college name') or data.get('college') or ''
                    room = data.get('room') or ''
                    
                    # Special fields that are also in the model
                    camera_data['gateway'] = gateway
                    camera_data['subnetMask'] = subnet
                    camera_data['macAddress'] = mac
                    camera_data['divisionName'] = college
                    camera_data['room'] = room
                    
                    # Sync to Master Location Registry
                    sync_location_to_master(college, camera_data['block'], camera_data['floor'], room, camera_data['brand'])
                    
                    # cameraId will be auto-generated by the model's save method if not provided
                    serial = data.get('serial number') or data.get('cameraid') or data.get('device serial number')
                    
                    if serial:
                        # Update existing or create new with this serial
                        obj, created = Camera.objects.update_or_create(
                            cameraId=serial,
                            defaults=camera_data
                        )
                        if created: created_count += 1
                        else: updated_count += 1
                    else:
                        # Create new, let save() handle sequential cameraId generation
                        Camera.objects.create(**camera_data)
                        created_count += 1
            
            log_activity(request.user, 'UPLOAD', 'Cameras', f"Uploaded file {file_obj.name}: {created_count} created, {updated_count} updated", request)
            return Response({
                'message': 'Import Complete',
                'created': created_count,
                'updated': updated_count,
                'total': len(rows_data)
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({'message': f'Error processing file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def add_remark(self, request, pk=None):
        camera = self.get_object()
        remark_text = request.data.get('remark')
        if not remark_text:
            return Response({'message': 'Remark is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        remark = CameraRemark.objects.create(
            camera=camera,
            remark=remark_text,
            device_status=camera.status,
            user=request.user
        )
        
        log_activity(request.user, 'REMARK', 'Cameras', f"Added remark to camera: {camera.name}", request)
        return Response(CameraRemarkSerializer(remark).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_relocation(self, request, pk=None):
        camera = self.get_object()
        serializer = CameraRelocationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(camera=camera, user=request.user)
        log_activity(request.user, 'RELOCATE', 'Cameras', f"Relocated camera: {camera.name}", request)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Cameras'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        instance = self.get_object()
        old_status = instance.status
        
        response = super().update(request, *args, **kwargs)
        
        if response.status_code == status.HTTP_200_OK:
            instance.refresh_from_db()
            new_status = instance.status
            if old_status != new_status:
                log_activity(request.user, 'STATUS_CHANGE', 'Cameras', f"Status changed from {old_status} to {new_status} for camera: {instance.name}", request)
                CameraRemark.objects.create(
                    camera=instance,
                    remark=f"System: Status changed from {old_status} to {new_status}",
                    device_status=new_status,
                    user=request.user
                )
            else:
                log_activity(request.user, 'EDIT', 'Cameras', f"Updated camera: {instance.name}", request)
        
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Cameras'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        log_activity(request.user, 'DELETE', 'Cameras', f"Deleted camera ID: {kwargs.get('pk')}", request)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='move-history')
    def move_history(self, request):
        from .models import CameraRelocation, NVRRelocation, BiometricRelocation, SwitchRelocation, RackRelocation
        history = []
        for r in CameraRelocation.objects.select_related('camera', 'user').all():
            history.append({
                'id': f"cam_{r.id}", 'dbId': r.camera.id,
                'deviceId': r.camera.cameraId, 'deviceName': r.camera.name,
                'deviceType': 'Camera', 'remark': r.remark,
                'timestamp': r.createdAt, 'user': r.user.name if r.user else 'System'
            })
        for r in NVRRelocation.objects.select_related('nvr', 'user').all():
            history.append({
                'id': f"nvr_{r.id}", 'dbId': r.nvr.id,
                'deviceId': r.nvr.serialNumber, 'deviceName': r.nvr.nvrName,
                'deviceType': 'NVR', 'remark': r.remark,
                'timestamp': r.createdAt, 'user': r.user.name if r.user else 'System'
            })
        for r in BiometricRelocation.objects.select_related('biometric', 'user').all():
            history.append({
                'id': f"bio_{r.id}", 'dbId': r.biometric.id,
                'deviceId': r.biometric.serialNumber, 'deviceName': r.biometric.name,
                'deviceType': 'Biometric', 'remark': r.remark,
                'timestamp': r.createdAt, 'user': r.user.name if r.user else 'System'
            })
        for r in SwitchRelocation.objects.select_related('switch', 'user').all():
            history.append({
                'id': f"sw_{r.id}", 'dbId': r.switch.id,
                'deviceId': r.switch.serialNumber, 'deviceName': r.switch.name,
                'deviceType': 'Switch', 'remark': r.remark,
                'timestamp': r.createdAt, 'user': r.user.name if r.user else 'System'
            })
        for r in RackRelocation.objects.select_related('rack', 'user').all():
            history.append({
                'id': f"rack_{r.id}", 'dbId': r.rack.id,
                'deviceId': r.rack.serialNumber, 'deviceName': r.rack.name,
                'deviceType': 'Rack', 'remark': r.remark,
                'timestamp': r.createdAt, 'user': r.user.name if r.user else 'System'
            })
        history.sort(key=lambda x: x['timestamp'], reverse=True)
        return Response(history)

class NVRViewSet(viewsets.ModelViewSet):
    queryset = NVR.objects.all().order_by('-id')
    serializer_class = NVRSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Storage'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        response = super().create(request, *args, **kwargs)
        log_activity(request.user, 'CREATE', 'NVR', f"Created NVR: {request.data.get('nvrName')}", request)
        return response

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Storage'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'message': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            if file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls'):
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
                rows_data = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
            elif file_obj.name.endswith('.csv'):
                import csv
                decoded_file = file_obj.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                rows_data = [{k.lower(): v for k, v in row.items()} for row in reader]
            else:
                return Response({'message': 'Unsupported format'}, status=status.HTTP_400_BAD_REQUEST)

            created_count = 0
            updated_count = 0
            with transaction.atomic():
                for data in rows_data:
                    college = data.get('college name') or data.get('college') or ''
                    block = data.get('block') or ''
                    floor = data.get('floor') or ''
                    room = data.get('room') or ''
                    
                    nvr_data = {
                        'ipAddress': data.get('ip address') or data.get('ipaddress'),
                        'nvrName': data.get('nvr name') or data.get('name') or 'Unknown NVR',
                        'location': data.get('location') or f"{college} | {block} | {floor} | {room}",
                        'brand': data.get('brand'),
                        'hardDisk': data.get('hard disk') or data.get('harddisk'),
                        'channel': data.get('channel'),
                        'status': data.get('status') or 'Online',
                        'divisionName': college,
                        'block': block,
                        'floor': floor,
                        'room': room,
                    }
                    
                    # Sync to Master Location Registry
                    sync_location_to_master(college, block, floor, room, nvr_data['brand'])
                    serial = data.get('serial number') or data.get('serialnumber') or data.get('sno')
                    
                    if serial:
                        obj, created = NVR.objects.update_or_create(serialNumber=serial, defaults=nvr_data)
                        if created: created_count += 1
                        else: updated_count += 1
                    else:
                        NVR.objects.create(**nvr_data)
                        created_count += 1
            
            log_activity(request.user, 'UPLOAD', 'NVR', f"Uploaded {file_obj.name}", request)
            return Response({'message': 'Import Complete', 'created': created_count, 'updated': updated_count}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Storage'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        data = request.data
        if not isinstance(data, list):
            return Response({'message': 'Data must be a list'}, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = self.get_serializer(data=data, many=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        log_activity(request.user, 'CREATE', 'NVR', f"Bulk created {len(data)} NVRs", request)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_remark(self, request, pk=None):
        nvr = self.get_object()
        remark_text = request.data.get('remark')
        if not remark_text:
            return Response({'message': 'Remark is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        remark = NVRRemark.objects.create(
            nvr=nvr,
            remark=remark_text,
            device_status=nvr.status,
            user=request.user
        )
        
        log_activity(request.user, 'REMARK', 'NVR', f"Added remark to NVR: {nvr.nvrName}", request)
        return Response(NVRRemarkSerializer(remark).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_relocation(self, request, pk=None):
        nvr = self.get_object()
        serializer = NVRRelocationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(nvr=nvr, user=request.user)
        log_activity(request.user, 'RELOCATE', 'NVR', f"Relocated NVR: {nvr.nvrName}", request)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Storage'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
            
        instance = self.get_object()
        old_status = instance.status
        
        response = super().update(request, *args, **kwargs)
        
        if response.status_code == status.HTTP_200_OK:
            instance.refresh_from_db()
            new_status = instance.status
            if old_status != new_status:
                log_activity(request.user, 'STATUS_CHANGE', 'NVR', f"Status changed from {old_status} to {new_status} for NVR: {instance.nvrName}", request)
                NVRRemark.objects.create(
                    nvr=instance,
                    remark=f"System: Status changed from {old_status} to {new_status}",
                    device_status=new_status,
                    user=request.user
                )
            else:
                log_activity(request.user, 'EDIT', 'NVR', f"Updated NVR: {instance.nvrName}", request)
                
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Storage'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        log_activity(request.user, 'DELETE', 'NVR', f"Deleted NVR ID: {kwargs.get('pk')}", request)
        return super().destroy(request, *args, **kwargs)

class BiometricViewSet(viewsets.ModelViewSet):
    queryset = Biometric.objects.all().order_by('-id')
    serializer_class = BiometricSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Biometrics'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        response = super().create(request, *args, **kwargs)
        log_activity(request.user, 'CREATE', 'Biometrics', f"Created Biometric: {request.data.get('name')}", request)
        return response

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Biometrics'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'message': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            if file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls'):
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
                rows_data = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
            elif file_obj.name.endswith('.csv'):
                import csv
                decoded_file = file_obj.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                rows_data = [{k.lower(): v for k, v in row.items()} for row in reader]
            else:
                return Response({'message': 'Unsupported format'}, status=status.HTTP_400_BAD_REQUEST)

            created_count = 0
            updated_count = 0
            with transaction.atomic():
                for data in rows_data:
                    college = data.get('college name') or data.get('college') or ''
                    block = data.get('block') or ''
                    floor = data.get('floor') or ''
                    room = data.get('room') or ''
                    
                    bio_data = {
                        'name': data.get('name') or 'Unknown Biometric',
                        'location': data.get('location') or f"{college} | {block} | {floor} | {room}",
                        'type': data.get('type') or 'Fingerprint',
                        'brand': data.get('brand'),
                        'ipAddress': data.get('ip address') or data.get('ipaddress'),
                        'serverIp': data.get('server ip') or data.get('serverip'),
                        'status': data.get('status') or 'Online',
                        'divisionName': college,
                        'block': block,
                        'floor': floor,
                        'room': room,
                    }
                    
                    # Sync to Master Location Registry
                    sync_location_to_master(college, block, floor, room, bio_data['brand'])
                    serial = data.get('serial number') or data.get('serialnumber')
                    
                    if serial:
                        obj, created = Biometric.objects.update_or_create(serialNumber=serial, defaults=bio_data)
                        if created: created_count += 1
                        else: updated_count += 1
                    else:
                        Biometric.objects.create(**bio_data)
                        created_count += 1
            
            log_activity(request.user, 'UPLOAD', 'Biometrics', f"Uploaded {file_obj.name}", request)
            return Response({'message': 'Import Complete', 'created': created_count, 'updated': updated_count}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Biometrics'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        data = request.data
        if not isinstance(data, list):
            return Response({'message': 'Data must be a list'}, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = self.get_serializer(data=data, many=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        log_activity(request.user, 'CREATE', 'Biometrics', f"Bulk created {len(data)} Biometric devices", request)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_remark(self, request, pk=None):
        biometric = self.get_object()
        remark_text = request.data.get('remark')
        if not remark_text:
            return Response({'message': 'Remark is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        remark = BiometricRemark.objects.create(
            biometric=biometric,
            remark=remark_text,
            device_status=biometric.status,
            user=request.user
        )
        
        log_activity(request.user, 'REMARK', 'Biometrics', f"Added remark to Biometric: {biometric.name}", request)
        return Response(BiometricRemarkSerializer(remark).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_relocation(self, request, pk=None):
        biometric = self.get_object()
        serializer = BiometricRelocationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(biometric=biometric, user=request.user)
        log_activity(request.user, 'RELOCATE', 'Biometrics', f"Relocated Biometric: {biometric.name}", request)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Biometrics'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
            
        instance = self.get_object()
        old_status = instance.status
        
        response = super().update(request, *args, **kwargs)
        
        if response.status_code == status.HTTP_200_OK:
            instance.refresh_from_db()
            new_status = instance.status
            if old_status != new_status:
                log_activity(request.user, 'STATUS_CHANGE', 'Biometrics', f"Status changed from {old_status} to {new_status} for Biometric: {instance.name}", request)
                BiometricRemark.objects.create(
                    biometric=instance,
                    remark=f"System: Status changed from {old_status} to {new_status}",
                    device_status=new_status,
                    user=request.user
                )
            else:
                log_activity(request.user, 'EDIT', 'Biometrics', f"Updated Biometric: {instance.name}", request)
                
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Biometrics'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        log_activity(request.user, 'DELETE', 'Biometrics', f"Deleted Biometric ID: {kwargs.get('pk')}", request)
        return super().destroy(request, *args, **kwargs)

class BarrierViewSet(viewsets.ModelViewSet):
    queryset = Barrier.objects.all().order_by('-id')
    serializer_class = BarrierSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin']:
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin']:
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin']:
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

class NetworkSwitchViewSet(viewsets.ModelViewSet):
    queryset = NetworkSwitch.objects.all().order_by('-id')
    serializer_class = NetworkSwitchSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Network Switches'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        response = super().create(request, *args, **kwargs)
        log_activity(request.user, 'CREATE', 'Switches', f"Created Switch: {request.data.get('name')}", request)
        return response

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Network Switches'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'message': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            if file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls'):
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
                rows_data = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
            elif file_obj.name.endswith('.csv'):
                import csv
                decoded_file = file_obj.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                rows_data = [{k.lower(): v for k, v in row.items()} for row in reader]
            else:
                return Response({'message': 'Unsupported format'}, status=status.HTTP_400_BAD_REQUEST)

            created_count = 0
            updated_count = 0
            with transaction.atomic():
                for data in rows_data:
                    college = data.get('college name') or data.get('college') or ''
                    block = data.get('block') or ''
                    floor = data.get('floor') or ''
                    room = data.get('room') or ''
                    
                    switch_data = {
                        'name': data.get('name') or 'Unknown Switch',
                        'ipAddress': data.get('ip address') or data.get('ipaddress'),
                        'location': data.get('location') or f"{college} | {block} | {floor} | {room}",
                        'brand': data.get('brand'),
                        'model': data.get('model'),
                        'portCount': data.get('port count') or data.get('portcount'),
                        'status': data.get('status') or 'Online',
                        'divisionName': college,
                        'block': block,
                        'floor': floor,
                        'room': room,
                    }
                    
                    # Sync to Master Location Registry
                    sync_location_to_master(college, block, floor, room, switch_data['brand'])
                    serial = data.get('serial number') or data.get('serialnumber')
                    
                    if serial:
                        obj, created = NetworkSwitch.objects.update_or_create(serialNumber=serial, defaults=switch_data)
                        if created: created_count += 1
                        else: updated_count += 1
                    else:
                        NetworkSwitch.objects.create(**switch_data)
                        created_count += 1
            
            log_activity(request.user, 'UPLOAD', 'Switches', f"Uploaded {file_obj.name}", request)
            return Response({'message': 'Import Complete', 'created': created_count, 'updated': updated_count}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Network Switches'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        data = request.data
        if not isinstance(data, list):
            return Response({'message': 'Data must be a list'}, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = self.get_serializer(data=data, many=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        log_activity(request.user, 'CREATE', 'Network Switches', f"Bulk created {len(data)} switches", request)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_remark(self, request, pk=None):
        switch = self.get_object()
        remark_text = request.data.get('remark')
        if not remark_text:
            return Response({'message': 'Remark is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        remark = SwitchRemark.objects.create(
            switch=switch,
            remark=remark_text,
            device_status=switch.status,
            user=request.user
        )
        
        log_activity(request.user, 'REMARK', 'Network Switches', f"Added remark to Switch: {switch.name}", request)
        return Response(SwitchRemarkSerializer(remark).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_relocation(self, request, pk=None):
        switch = self.get_object()
        serializer = SwitchRelocationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(switch=switch, user=request.user)
        log_activity(request.user, 'RELOCATE', 'Network Switches', f"Relocated Switch: {switch.name}", request)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Network Switches'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
            
        instance = self.get_object()
        old_status = instance.status
        
        response = super().update(request, *args, **kwargs)
        
        if response.status_code == status.HTTP_200_OK:
            instance.refresh_from_db()
            new_status = instance.status
            if old_status != new_status:
                log_activity(request.user, 'STATUS_CHANGE', 'Switches', f"Status changed from {old_status} to {new_status} for Switch: {instance.name}", request)
                SwitchRemark.objects.create(
                    switch=instance,
                    remark=f"System: Status changed from {old_status} to {new_status}",
                    device_status=new_status,
                    user=request.user
                )
            else:
                log_activity(request.user, 'EDIT', 'Switches', f"Updated Switch: {instance.name}", request)
                
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Network Switches'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        log_activity(request.user, 'DELETE', 'Switches', f"Deleted Switch ID: {kwargs.get('pk')}", request)
        return super().destroy(request, *args, **kwargs)

class CameraRemarkViewSet(viewsets.ModelViewSet):
    queryset = CameraRemark.objects.all()
    serializer_class = CameraRemarkSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class NVRRemarkViewSet(viewsets.ModelViewSet):
    queryset = NVRRemark.objects.all()
    serializer_class = NVRRemarkSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class BiometricRemarkViewSet(viewsets.ModelViewSet):
    queryset = BiometricRemark.objects.all()
    serializer_class = BiometricRemarkSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class SwitchRemarkViewSet(viewsets.ModelViewSet):
    queryset = SwitchRemark.objects.all()
    serializer_class = SwitchRemarkSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class ActivityLogViewSet(viewsets.ModelViewSet):
    serializer_class = ActivityLogSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'Super Admin':
            return ActivityLog.objects.all().order_by('-timestamp')
        return ActivityLog.objects.filter(user=user).order_by('-timestamp')

    def create(self, request, *args, **kwargs):
        action = request.data.get('action', 'VIEW')
        page = request.data.get('page', 'Unknown')
        details = request.data.get('details', '')
        log_activity(request.user, action, page, details, request)
        return Response({'status': 'logged'})

class LocationViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        from .models import Block, Floor, Room, GlobalSiteConfig
        
        default_division = ''
        default_assignee = None
        try:
            site = GlobalSiteConfig.objects.first()
            if not site:
                from users.models import User
                default_user = User.objects.filter(name='Karthik K N').first() or User.objects.filter(role='Super Admin').first() or User.objects.first()
                if default_user:
                    site = GlobalSiteConfig.objects.create(
                        divisionName='RGI',
                        assignedTo=default_user
                    )
            if site:
                default_division = site.divisionName or ''
                if site.assignedTo:
                    from users.serializers import UserSerializer
                    default_assignee = UserSerializer(site.assignedTo).data
        except Exception as e:
            print("Error initializing GlobalSiteConfig:", e)

        locations = []
        blocks = Block.objects.all()
        idx = 1
        
        for b in blocks:
            floors = Floor.objects.filter(block=b)
            if not floors.exists():
                locations.append({ 'id': f"b-{b.id}", 'divisionName': default_division, 'block': b.name, 'floor': '', 'room': '', 'assignedTo': default_assignee })
            else:
                for f in floors:
                    rooms = Room.objects.filter(floor=f)
                    if not rooms.exists():
                        locations.append({ 'id': f"f-{f.id}", 'divisionName': default_division, 'block': b.name, 'floor': f.name, 'room': '', 'assignedTo': default_assignee })
                    else:
                        for r in rooms:
                            locations.append({ 'id': f"r-{r.id}", 'divisionName': default_division, 'block': b.name, 'floor': f.name, 'room': r.name, 'assignedTo': default_assignee })
                            
        return Response(locations)

    def create(self, request):
        from .models import Block, Floor, Room
        block_name = request.data.get('block', '').strip().upper()
        floor_name = request.data.get('floor', '').strip().upper()
        room_name = request.data.get('room', '').strip().upper()
        
        if not block_name:
            return Response({'message': 'Block name is required'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            b, b_created = Block.objects.get_or_create(name=block_name)
            if not b_created and not floor_name and not room_name:
                return Response({'message': f'Duplicate Block: {block_name} already exists.'}, status=status.HTTP_400_BAD_REQUEST)
                
            msg = f"Added block {block_name}"
            
            if floor_name:
                f, f_created = Floor.objects.get_or_create(name=floor_name, block=b)
                if not f_created and not room_name:
                    return Response({'message': f'Duplicate Floor: {floor_name} already exists in {block_name}.'}, status=status.HTTP_400_BAD_REQUEST)
                msg += f", floor {floor_name}"
                if room_name:
                    r, r_created = Room.objects.get_or_create(name=room_name, block=b, floor=f)
                    if not r_created:
                        return Response({'message': f'Duplicate Room: {room_name} already exists.'}, status=status.HTTP_400_BAD_REQUEST)
                    msg += f", room {room_name}"
            
            log_activity(request.user, 'CREATE', 'Location', msg, request)
            return Response({'message': 'Location saved'})
        except Exception as e:
            log_activity(request.user, 'ERROR', 'Location', f"Save failed: {str(e)}", request)
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, pk=None):
        from .models import Block, Floor, Room, Camera, NVR, Biometric, NetworkSwitch
        # pk format: b-1, f-2, r-3 or legacy-...
        if not pk:
            return Response(status=status.HTTP_400_BAD_REQUEST)
            
        parts = pk.split('-')
        prefix = parts[0]
        db_id = parts[1] if len(parts) > 1 else None

        if prefix == 'legacy':
            return Response({'message': 'Cannot delete legacy locations. Please reassign the devices first.'}, status=status.HTTP_400_BAD_REQUEST)

        # Check for devices in this location
        def has_devices(block_name=None, floor_name=None, room_name=None):
            for model in [Camera, NVR, Biometric, NetworkSwitch]:
                qs = model.objects.all()
                if block_name: qs = qs.filter(block__iexact=block_name)
                if floor_name: qs = qs.filter(floor__iexact=floor_name)
                if room_name: qs = qs.filter(room__iexact=room_name)
                if qs.exists(): return True
            return False

        if prefix == 'b':
            block = Block.objects.filter(id=db_id).first()
            if block and has_devices(block_name=block.name):
                return Response({'message': f'Cannot delete Block {block.name}: Devices are still assigned to it.'}, status=status.HTTP_400_BAD_REQUEST)
            if block: block.delete()
        elif prefix == 'f':
            floor = Floor.objects.filter(id=db_id).first()
            if floor and has_devices(block_name=floor.block.name, floor_name=floor.name):
                return Response({'message': f'Cannot delete Floor {floor.name}: Devices are still assigned to it.'}, status=status.HTTP_400_BAD_REQUEST)
            if floor: floor.delete()
        elif prefix == 'r':
            room = Room.objects.filter(id=db_id).first()
            if room and has_devices(block_name=room.block.name, floor_name=room.floor.name, room_name=room.name):
                return Response({'message': f'Cannot delete Room {room.name}: Devices are still assigned to it.'}, status=status.HTTP_400_BAD_REQUEST)
            if room: room.delete()
            
        return Response({'message': 'Location deleted'})

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        # We restore the CSV upload functionality using native tables
        if request.user.role not in ['Super Admin', 'Admin']:
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'message': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            if file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
                rows_data = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
            elif file_obj.name.endswith('.csv'):
                import csv
                decoded_file = file_obj.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                rows_data = [{k.lower(): v for k, v in row.items()} for row in reader]
            else:
                return Response({'message': 'Unsupported format'}, status=status.HTTP_400_BAD_REQUEST)

            from .models import Block, Floor, Room
            created_count = 0
            skipped_count = 0
            with transaction.atomic():
                for data in rows_data:
                    block = (data.get('block') or '').strip().upper()
                    floor = (data.get('floor') or '').strip().upper()
                    room = (data.get('room') or '').strip().upper()
                    
                    if not block: continue
                    
                    b, created_b = Block.objects.get_or_create(name=block)
                    if created_b: 
                        created_count += 1
                    else:
                        skipped_count += 1
                        
                    if floor:
                        f, created_f = Floor.objects.get_or_create(name=floor, block=b)
                        if created_f: 
                            created_count += 1
                        else:
                            skipped_count += 1
                            
                        if room:
                            r, created_r = Room.objects.get_or_create(name=room, block=b, floor=f)
                            if created_r: 
                                created_count += 1
                            else:
                                skipped_count += 1
            
            log_message = f"Uploaded {file_obj.name}. Created {created_count} new entries."
            if skipped_count > 0:
                log_message += f" Skipped {skipped_count} duplicate entries."
                
            log_activity(request.user, 'UPLOAD', 'Location', log_message, request)
            return Response({'message': 'Import Complete', 'created': created_count, 'skipped': skipped_count}, status=status.HTTP_201_CREATED)
        except Exception as e:
            log_activity(request.user, 'ERROR', 'Location', f"Upload failed: {str(e)}", request)
            return Response({'message': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class GlobalSiteConfigViewSet(viewsets.ModelViewSet):
    queryset = GlobalSiteConfig.objects.all()
    serializer_class = GlobalSiteConfigSerializer
    permission_classes = [IsAuthenticated]

    def list(self, request):
        config = GlobalSiteConfig.objects.first()
        if not config:
            return Response({})
        return Response(GlobalSiteConfigSerializer(config).data)

    @action(detail=False, methods=['post'])
    def apply_site(self, request):
        if request.user.role not in ['Super Admin', 'Admin']:
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        
        config = GlobalSiteConfig.objects.first()
        if not config:
            config = GlobalSiteConfig.objects.create()
        
        serializer = self.get_serializer(config, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        
        log_activity(request.user, 'CONFIG_UPDATE', 'Global', f"Updated global site config: {config.divisionName}", request)
        
        return Response(serializer.data)


    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        from maintenance.models import Ticket
        
        return Response({
            'cameras': {
                'total': Camera.objects.count(),
                'online': Camera.objects.filter(status='Online').count(),
                'offline': Camera.objects.filter(status='Offline').count(),
                'maintenance': Camera.objects.filter(status='Maintenance').count(),
                'scrap': Camera.objects.filter(status='Scrap').count(),
                'official': Camera.objects.count(),
            },
            'nvrs': {
                'total': NVR.objects.count(),
                'online': NVR.objects.filter(status='Online').count(),
                'offline': NVR.objects.filter(status='Offline').count(),
                'maintenance': NVR.objects.filter(status='Maintenance').count(),
                'scrap': NVR.objects.filter(status='Scrap').count(),
                'official': NVR.objects.count(),
            },
            'biometrics': {
                'total': Biometric.objects.count(),
                'online': Biometric.objects.filter(status='Online').count(),
                'offline': Biometric.objects.filter(status='Offline').count(),
                'maintenance': Biometric.objects.filter(status='Maintenance').count(),
                'scrap': Biometric.objects.filter(status='Scrap').count(),
                'official': Biometric.objects.count(),
            },
            'switches': {
                'total': NetworkSwitch.objects.count(),
                'online': NetworkSwitch.objects.filter(status='Online').count(),
                'offline': NetworkSwitch.objects.filter(status='Offline').count(),
                'maintenance': NetworkSwitch.objects.filter(status='Maintenance').count(),
                'scrap': NetworkSwitch.objects.filter(status='Scrap').count(),
                'official': NetworkSwitch.objects.count(),
            },
            'tickets': {
                'total': Ticket.objects.count(),
                'open': Ticket.objects.filter(status='Open').count(),
                'inProgress': Ticket.objects.filter(status='In Progress').count(),
                'completed': Ticket.objects.filter(status='Completed').count(),
                'upgrade': Ticket.objects.filter(category__icontains='Upgrade').count(),
                'project': Ticket.objects.filter(category__icontains='Project').count(),
            },
            'distribution': [
                {'name': c['divisionName'] or 'Unassigned', 'count': c['count']}
                for c in Camera.objects.values('divisionName').annotate(count=Count('id'))
            ]
        })

class RackViewSet(viewsets.ModelViewSet):
    queryset = Rack.objects.all().order_by('-id')
    serializer_class = RackSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Racks'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        response = super().create(request, *args, **kwargs)
        rack_name = request.data.get('name', 'Unknown')
        serial_no = request.data.get('serialNumber', '')
        details_str = f"Created Rack: {rack_name}"
        if serial_no: details_str += f" (Asset: {serial_no})"
        log_activity(request.user, 'CREATE', 'Racks', details_str, request)
        return response

    @action(detail=True, methods=['post'])
    def add_remark(self, request, pk=None):
        rack = self.get_object()
        remark_text = request.data.get('remark')
        if not remark_text:
            return Response({'message': 'Remark is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        remark = RackRemark.objects.create(
            rack=rack,
            remark=remark_text,
            device_status=rack.status,
            user=request.user
        )
        
        log_activity(request.user, 'REMARK', 'Racks', f"Added remark to Rack: {rack.name}", request)
        return Response(RackRemarkSerializer(remark).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def add_relocation(self, request, pk=None):
        rack = self.get_object()
        serializer = RackRelocationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(rack=rack, user=request.user)
        log_activity(request.user, 'RELOCATE', 'Racks', f"Relocated Rack: {rack.name}", request)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Racks'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
            
        instance = self.get_object()
        old_status = instance.status
        
        response = super().update(request, *args, **kwargs)
        
        if response.status_code == status.HTTP_200_OK:
            instance.refresh_from_db()
            new_status = instance.status
            if old_status != new_status:
                log_activity(request.user, 'STATUS_CHANGE', 'Racks', f"Status changed from {old_status} to {new_status} for Rack: {instance.name}", request)
                RackRemark.objects.create(
                    rack=instance,
                    remark=f"System: Status changed from {old_status} to {new_status}",
                    device_status=new_status,
                    user=request.user
                )
            else:
                details_str = f"Updated Rack: {instance.name}"
                if instance.serialNumber: details_str += f" (Asset: {instance.serialNumber})"
                log_activity(request.user, 'EDIT', 'Racks', details_str, request)
                
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ['Super Admin', 'Admin'] and not check_can_edit(request.user, 'Racks'):
            return Response({'message': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)
        log_activity(request.user, 'DELETE', 'Racks', f"Deleted Rack ID: {kwargs.get('pk')}", request)
        return super().destroy(request, *args, **kwargs)

class RackRemarkViewSet(viewsets.ModelViewSet):
    queryset = RackRemark.objects.all()
    serializer_class = RackRemarkSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class DivisionViewSet(viewsets.ModelViewSet):
    queryset = Division.objects.all().order_by('-createdAt')
    serializer_class = DivisionSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
        if 'name' in data and data['name']:
            data['name'] = str(data['name']).strip().upper()
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
        if 'name' in data and data['name']:
            data['name'] = str(data['name']).strip().upper()
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def perform_create(self, serializer):
        serializer.save()
        log_activity(self.request.user, 'CREATE', 'Division', f"Created division: {self.request.data.get('name')}", self.request)

    def perform_update(self, serializer):
        instance = serializer.save()
        log_activity(self.request.user, 'EDIT', 'Division', f"Updated division: {instance.name}", self.request)

    def perform_destroy(self, instance):
        log_activity(self.request.user, 'DELETE', 'Division', f"Deleted division: {instance.name}", self.request)
        instance.delete()

    @action(detail=False, methods=['post'])
    def merge(self, request):
        old_names = request.data.get('old_names', [])
        new_name = request.data.get('new_name')
        division_type = request.data.get('division_type', 'College')

        if not old_names or not new_name:
            return Response({"error": "old_names and new_name are required."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            # Create or update the master division
            division, created = Division.objects.get_or_create(
                name=new_name,
                defaults={'division_type': division_type}
            )
            
            # Store merged legacy names
            existing_merged = division.merged_from if division.merged_from else []
            merged_set = set(existing_merged)
            merged_set.update(old_names)
            division.merged_from = list(merged_set)
            division.save()

            # Update all standard CCTV devices
            Camera.objects.filter(divisionName__in=old_names).update(divisionName=new_name)
            NVR.objects.filter(divisionName__in=old_names).update(divisionName=new_name)
            Biometric.objects.filter(divisionName__in=old_names).update(divisionName=new_name)
            NetworkSwitch.objects.filter(divisionName__in=old_names).update(divisionName=new_name)
            Rack.objects.filter(divisionName__in=old_names).update(divisionName=new_name)

            # Update Tickets in maintenance app
            try:
                from maintenance.models import Ticket
                Ticket.objects.filter(divisionName__in=old_names).update(divisionName=new_name)
            except ImportError:
                pass # Maintenance app might not be installed or available

            # (User requested to KEEP the old divisions A and B in the registry after merging)
            # Division.objects.filter(name__in=old_names).exclude(name=new_name).delete()

        log_activity(request.user, 'MERGE', 'Division', f"Merged {len(old_names)} legacy colleges into {new_name}", request)
        return Response({"message": "Successfully merged colleges.", "new_name": new_name}, status=status.HTTP_200_OK)

class BrandViewSet(viewsets.ModelViewSet):
    queryset = Brand.objects.all().order_by('-createdAt')
    serializer_class = BrandSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
        if 'name' in data and data['name']:
            data['name'] = str(data['name']).strip().upper()
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
        if 'name' in data and data['name']:
            data['name'] = str(data['name']).strip().upper()
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def perform_create(self, serializer):
        serializer.save()
        log_activity(self.request.user, 'CREATE', 'Brand', f"Created brand: {self.request.data.get('name')}", self.request)

    def perform_update(self, serializer):
        instance = serializer.save()
        log_activity(self.request.user, 'EDIT', 'Brand', f"Updated brand: {instance.name}", self.request)

    def perform_destroy(self, instance):
        log_activity(self.request.user, 'DELETE', 'Brand', f"Deleted brand: {instance.name}", self.request)
        instance.delete()
